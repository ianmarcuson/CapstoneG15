"""
generar_intervalos.py

Genera X replicas de DatosV2.xlsx cambiando SOLO las simulaciones de llegadas
Poisson en la hoja Motor_Arribos, y luego ejecuta model_interdia.py para cada
replica, dejando outputs atribuibles a cada una en una subcarpeta.

Uso tipico:
    python generar_intervalos.py --replicas 10 --workers 2 --overwrite
    python generar_intervalos.py --start-replica 11 --replicas 20 --time-limit 250 --workers 4 --overwrite

Por defecto guarda todo en:
    ./resultados_intervalos/

Supuestos:
    - Este archivo esta en el mismo directorio que:
        DatosV2.xlsx
        model_interdia.py
        params.py
    - El modelo interdia lee params.EXCEL_PATH.
    - Las salidas base del modelo son:
        solution_interday.xlsx
        schedule_resultado.csv
        resumen_diario.csv

El script NO modifica model_interdia.py ni params.py en disco. Para cada corrida,
lanza un proceso de Python que parchea params en memoria y luego ejecuta
model_interdia.main(). Además fuerza UTF-8 en los subprocess para evitar
errores UnicodeEncodeError en Windows al escribir logs.
"""

from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook


# =============================================================================
# Utilidades de Excel
# =============================================================================

def _header_map(ws, row: int = 1) -> Dict[str, int]:
    """Retorna {nombre_columna: indice_columna_1_based}."""
    out = {}
    for cell in ws[row]:
        if cell.value is not None:
            out[str(cell.value).strip()] = cell.column
    return out


def _get_required(mapping: Dict[str, int], name: str, sheet_name: str) -> int:
    if name not in mapping:
        raise ValueError(
            f"No se encontro la columna '{name}' en la hoja '{sheet_name}'. "
            f"Columnas disponibles: {list(mapping.keys())}"
        )
    return mapping[name]


def leer_tasas_configuracion(xlsx_path: Path) -> Dict[int, float]:
    """Lee tasas desde Configuracion: columna Id y columna Tasa de Llegada."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    if "Configuracion" not in wb.sheetnames:
        raise ValueError("El archivo no tiene hoja 'Configuracion'.")

    ws = wb["Configuracion"]
    headers = _header_map(ws)
    id_col = _get_required(headers, "Id", "Configuracion")
    tasa_col = _get_required(headers, "Tasa de Llegada", "Configuracion")

    tasas: Dict[int, float] = {}
    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, id_col).value
        if raw_id is None or raw_id == "":
            continue
        try:
            tipo_id = int(raw_id)
        except (TypeError, ValueError):
            continue

        raw_tasa = ws.cell(r, tasa_col).value
        if raw_tasa is None or raw_tasa == "":
            tasa = 0.0
        else:
            tasa = float(raw_tasa)

        if tasa < 0:
            raise ValueError(f"Tasa negativa para tipo {tipo_id}: {tasa}")

        tasas[tipo_id] = tasa

    if not tasas:
        raise ValueError("No se encontraron tasas en Configuracion.")

    wb.close()
    return tasas


def columnas_tipos_motor_arribos(ws) -> List[Tuple[int, int]]:
    """
    Retorna lista de pares (tipo_id, columna) para columnas 'Tipo 1', 'Tipo 2', etc.
    """
    out: List[Tuple[int, int]] = []
    for cell in ws[1]:
        value = cell.value
        if not isinstance(value, str):
            continue
        value = value.strip()
        if not value.lower().startswith("tipo"):
            continue
        try:
            tipo_id = int(value.split()[-1])
        except (IndexError, ValueError):
            continue
        out.append((tipo_id, cell.column))

    out.sort(key=lambda x: x[0])
    if not out:
        raise ValueError("No se encontraron columnas 'Tipo X' en Motor_Arribos.")
    return out


def filas_dias_motor_arribos(ws) -> List[int]:
    """Filas con un Dia valido en la columna A."""
    rows: List[int] = []
    for r in range(2, ws.max_row + 1):
        dia = ws.cell(r, 1).value
        if dia is None or dia == "":
            continue
        try:
            int(dia)
        except (TypeError, ValueError):
            continue
        rows.append(r)
    if not rows:
        raise ValueError("No se encontraron dias validos en Motor_Arribos.")
    return rows



def normalizar_duracion_configuracion_wb(wb) -> int:
    """
    Convierte Configuracion['Duracion (Dias)'] a valores numericos.

    Motivo: openpyxl preserva formulas, pero no calcula ni mantiene siempre
    el valor cacheado. Luego pandas puede leer esas formulas como NaN y
    model_interdia.py falla al hacer int(row['Duracion (Dias)']).

    Formula equivalente usada en el Excel:
        (Ciclos - 1) * ((Sesiones - 1) * TBS + TBC) + (Sesiones - 1) * TBS
    """
    if "Configuracion" not in wb.sheetnames:
        return 0

    ws = wb["Configuracion"]
    headers = _header_map(ws)

    required = ["Id", "Ciclos", "Sesiones", "TBS", "TBC", "Duracion (Dias)"]
    if any(name not in headers for name in required):
        return 0

    id_col = headers["Id"]
    ciclos_col = headers["Ciclos"]
    sesiones_col = headers["Sesiones"]
    tbs_col = headers["TBS"]
    tbc_col = headers["TBC"]
    dur_col = headers["Duracion (Dias)"]

    changed = 0
    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, id_col).value
        if raw_id is None or raw_id == "":
            continue

        try:
            ciclos = int(ws.cell(r, ciclos_col).value)
            sesiones = int(ws.cell(r, sesiones_col).value)
            tbs = int(ws.cell(r, tbs_col).value)
            tbc = int(ws.cell(r, tbc_col).value)
        except (TypeError, ValueError):
            continue

        duracion = (ciclos - 1) * ((sesiones - 1) * tbs + tbc) + (sesiones - 1) * tbs
        ws.cell(r, dur_col).value = int(duracion)
        changed += 1

    return changed


def reparar_duracion_configuracion(xlsx_path: Path) -> int:
    """Repara una replica existente dejando Duracion (Dias) como numero."""
    wb = load_workbook(xlsx_path, data_only=False)
    changed = normalizar_duracion_configuracion_wb(wb)
    if changed:
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
        except Exception:
            pass
        wb.save(xlsx_path)
    wb.close()
    return changed

def generar_replica_datosv2(
    base_xlsx: Path,
    out_xlsx: Path,
    seed: int,
    overwrite: bool = False,
) -> Dict[str, object]:
    """
    Crea una replica del Excel base, reemplazando solo Motor_Arribos!Tipo X
    por simulaciones Poisson con lambda igual a la tasa fija de Configuracion.
    """
    if out_xlsx.exists() and not overwrite:
        raise FileExistsError(f"Ya existe {out_xlsx}. Usa --overwrite para reemplazarlo.")

    tasas = leer_tasas_configuracion(base_xlsx)

    wb = load_workbook(base_xlsx, data_only=False)
    if "Motor_Arribos" not in wb.sheetnames:
        raise ValueError("El archivo no tiene hoja 'Motor_Arribos'.")

    ws_arribos = wb["Motor_Arribos"]
    tipo_cols = columnas_tipos_motor_arribos(ws_arribos)
    day_rows = filas_dias_motor_arribos(ws_arribos)

    # Validar que cada columna Tipo tenga tasa.
    missing_rates = [tipo_id for tipo_id, _ in tipo_cols if tipo_id not in tasas]
    if missing_rates:
        raise ValueError(f"Faltan tasas para tipos: {missing_rates}")

    rng = np.random.default_rng(seed)

    total_llegadas = 0
    for tipo_id, col in tipo_cols:
        lam = tasas[tipo_id]
        values = rng.poisson(lam=lam, size=len(day_rows))
        total_llegadas += int(values.sum())
        for row_idx, value in zip(day_rows, values):
            ws_arribos.cell(row=row_idx, column=col).value = int(value)

    # Evitar NaN en pandas: openpyxl no calcula formulas al guardar.
    normalizar_duracion_configuracion_wb(wb)

    # Dejar indicado a Excel que recalcule formulas al abrir, por si usan hojas derivadas.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    wb.close()

    return {
        "archivo": str(out_xlsx),
        "seed": seed,
        "n_dias": len(day_rows),
        "n_tipos": len(tipo_cols),
        "total_llegadas_simuladas": total_llegadas,
    }


# =============================================================================
# Ejecucion del modelo interdia por replica
# =============================================================================

@dataclass
class RunResult:
    replica: int
    seed: int
    datos_xlsx: str
    output_xlsx: str
    output_schedule_csv: str
    output_resumen_csv: str
    log_file: str
    returncode: int
    status: str
    elapsed_seconds: float


def build_model_runner_code(
    base_dir: Path,
    replica_xlsx: Path,
    output_xlsx: str,
    output_csv: str,
    output_summary_csv: str,
    gurobi_threads: Optional[int],
    time_limit: Optional[int],
    mip_gap: Optional[float],
    force_base_only: bool,
) -> str:
    """
    Codigo Python que se ejecuta en un subprocess. Parchea params en memoria.
    No modifica params.py en disco.
    """
    lines = [
        "from pathlib import Path",
        "import sys",
        "# Forzar salida UTF-8 en Windows aunque stdout/stderr vayan a archivo.",
        "try:",
        "    sys.stdout.reconfigure(encoding='utf-8', errors='replace')",
        "    sys.stderr.reconfigure(encoding='utf-8', errors='replace')",
        "except Exception:",
        "    pass",
        f"sys.path.insert(0, {str(base_dir)!r})",
        "import params as P",
        f"P.EXCEL_PATH = Path({str(replica_xlsx)!r})",
        f"P.OUTPUT_XLSX = {output_xlsx!r}",
        f"P.OUTPUT_CSV = {output_csv!r}",
        f"P.OUTPUT_SUMMARY_CSV = {output_summary_csv!r}",
    ]

    if gurobi_threads is not None:
        lines.append(f"P.THREADS = {int(gurobi_threads)}")
    if time_limit is not None:
        lines.append(f"P.TIME_LIMIT_SECONDS = {int(time_limit)}")
    if mip_gap is not None:
        lines.append(f"P.MIP_GAP = {float(mip_gap)!r}")
    if force_base_only:
        lines.append("P.RUN_ALL_SCENARIOS = False")

    lines.extend([
        "import model_interdia",
        "model_interdia.main()",
    ])
    return "\n".join(lines)


def run_model_for_replica(
    replica: int,
    seed: int,
    base_dir: Path,
    out_dir: Path,
    replica_xlsx: Path,
    python_exe: str,
    gurobi_threads: Optional[int],
    time_limit: Optional[int],
    mip_gap: Optional[float],
    force_base_only: bool,
) -> RunResult:
    output_xlsx_path = out_dir / f"solution_interday-{replica}.xlsx"
    output_csv_path = out_dir / f"schedule_resultado-{replica}.csv"
    output_summary_csv_path = out_dir / f"resumen_diario-{replica}.csv"
    log_file = out_dir / f"log_interdia-{replica}.txt"

    code = build_model_runner_code(
        base_dir=base_dir.resolve(),
        replica_xlsx=replica_xlsx.resolve(),
        output_xlsx=str(output_xlsx_path.resolve()),
        output_csv=str(output_csv_path.resolve()),
        output_summary_csv=str(output_summary_csv_path.resolve()),
        gurobi_threads=gurobi_threads,
        time_limit=time_limit,
        mip_gap=mip_gap,
        force_base_only=force_base_only,
    )

    t0 = time.time()
    with open(log_file, "w", encoding="utf-8") as log:
        log.write(f"Replica: {replica}\n")
        log.write(f"Seed: {seed}\n")
        log.write(f"Datos: {replica_xlsx}\n")
        log.write(f"Output XLSX: {output_xlsx_path}\n")
        log.write("\n--- INICIO MODELO ---\n\n")
        log.flush()

        env = os.environ.copy()
        env["PYTHONUTF8"] = "1"
        env["PYTHONIOENCODING"] = "utf-8"

        completed = subprocess.run(
            [python_exe, "-X", "utf8", "-c", code],
            cwd=out_dir,
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
        )

    elapsed = time.time() - t0
    status = "OK" if completed.returncode == 0 else "ERROR"

    return RunResult(
        replica=replica,
        seed=seed,
        datos_xlsx=str(replica_xlsx),
        output_xlsx=str(output_xlsx_path),
        output_schedule_csv=str(output_csv_path),
        output_resumen_csv=str(output_summary_csv_path),
        log_file=str(log_file),
        returncode=completed.returncode,
        status=status,
        elapsed_seconds=round(elapsed, 2),
    )


# =============================================================================
# Main
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera replicas Poisson de DatosV2 y ejecuta model_interdia para cada replica."
    )
    parser.add_argument("--replicas", type=int, default=10, help="Cantidad de replicas a generar/correr.")
    parser.add_argument("--start-replica", type=int, default=1, help="Numero inicial de replica. Ej: 11 genera/corre 11..(11+replicas-1).")
    parser.add_argument("--base-dir", type=str, default=".", help="Directorio donde estan DatosV2.xlsx, params.py y model_interdia.py.")
    parser.add_argument("--base-excel", type=str, default="DatosV2.xlsx", help="Archivo base Excel.")
    parser.add_argument("--out-dir", type=str, default="resultados_intervalos", help="Subcarpeta donde se guardan replicas, outputs, logs y resumen.")
    parser.add_argument("--seed", type=int, default=12345, help="Seed base. La replica i usa seed + i.")
    parser.add_argument("--workers", type=int, default=2, help="Cantidad de replicas a resolver en paralelo.")
    parser.add_argument("--gurobi-threads", type=int, default=1, help="Threads de Gurobi por replica. Usa 1 con workers=2.")
    parser.add_argument("--time-limit", type=int, default=None, help="Opcional: sobrescribe P.TIME_LIMIT_SECONDS.")
    parser.add_argument("--mip-gap", type=float, default=None, help="Opcional: sobrescribe P.MIP_GAP.")
    parser.add_argument("--overwrite", action="store_true", help="Sobrescribe replicas/outputs existentes.")
    parser.add_argument("--solo-generar", action="store_true", help="Solo genera DatosV2-i.xlsx; no corre model_interdia.")
    parser.add_argument("--solo-correr", action="store_true", help="No genera Excel; usa replicas DatosV2-i.xlsx ya existentes.")
    parser.add_argument("--keep-scenarios", action="store_true", help="No fuerza RUN_ALL_SCENARIOS=False.")
    return parser.parse_args()


def validar_archivos_base(base_dir: Path, base_excel: Path) -> None:
    required = [base_excel, base_dir / "params.py", base_dir / "model_interdia.py"]
    missing = [str(p) for p in required if not p.exists()]
    if missing:
        raise FileNotFoundError("Faltan archivos requeridos:\n  - " + "\n  - ".join(missing))


def write_summary_csv(path: Path, results: List[RunResult]) -> None:
    if not results:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(results[0]).keys()))
        writer.writeheader()
        for result in results:
            writer.writerow(asdict(result))


def main() -> None:
    args = parse_args()
    if args.replicas <= 0:
        raise ValueError("--replicas debe ser >= 1")
    if args.start_replica <= 0:
        raise ValueError("--start-replica debe ser >= 1")
    if args.workers <= 0:
        raise ValueError("--workers debe ser >= 1")
    if args.solo_generar and args.solo_correr:
        raise ValueError("No puedes usar --solo-generar y --solo-correr al mismo tiempo.")

    base_dir = Path(args.base_dir).resolve()
    base_excel = Path(args.base_excel)
    if not base_excel.is_absolute():
        base_excel = base_dir / base_excel

    validar_archivos_base(base_dir, base_excel)

    out_dir = Path(args.out_dir)
    if not out_dir.is_absolute():
        out_dir = base_dir / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 72)
    print("GENERADOR DE REPLICAS + EJECUCION MODEL_INTERDIA")
    print("=" * 72)
    print(f"Directorio modelo:{base_dir}")
    print(f"Carpeta salida:   {out_dir}")
    print(f"Excel base:       {base_excel}")
    end_replica = args.start_replica + args.replicas - 1
    print(f"Replicas:         {args.replicas} ({args.start_replica}..{end_replica})")
    print(f"Workers modelo:   {args.workers}")
    print(f"Threads/replica:  {args.gurobi_threads}")
    print("=" * 72)

    # -------------------------------------------------------------------------
    # 1) Generar replicas DatosV2-i.xlsx
    # -------------------------------------------------------------------------
    replica_paths: List[Tuple[int, int, Path]] = []
    if not args.solo_correr:
        for i in range(args.start_replica, end_replica + 1):
            seed_i = args.seed + i
            replica_path = out_dir / f"DatosV2-{i}.xlsx"
            info = generar_replica_datosv2(
                base_xlsx=base_excel,
                out_xlsx=replica_path,
                seed=seed_i,
                overwrite=args.overwrite,
            )
            replica_paths.append((i, seed_i, replica_path))
            print(
                f"[Replica {i}] generado {replica_path.name} | "
                f"seed={seed_i} | llegadas={info['total_llegadas_simuladas']}"
            )
    else:
        for i in range(args.start_replica, end_replica + 1):
            seed_i = args.seed + i
            replica_path = out_dir / f"DatosV2-{i}.xlsx"
            if not replica_path.exists():
                raise FileNotFoundError(f"No existe {replica_path}. Quita --solo-correr o genera antes.")
            # Reparar formulas no calculadas de Configuracion si la replica ya existia.
            reparar_duracion_configuracion(replica_path)
            replica_paths.append((i, seed_i, replica_path))

    if args.solo_generar:
        print("\nListo: replicas generadas. No se ejecuto model_interdia por --solo-generar.")
        return

    # -------------------------------------------------------------------------
    # 2) Ejecutar model_interdia para cada replica
    # -------------------------------------------------------------------------
    print("\nEjecutando model_interdia por replica...")
    results: List[RunResult] = []

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = []
        for i, seed_i, replica_path in replica_paths:
            futures.append(
                executor.submit(
                    run_model_for_replica,
                    replica=i,
                    seed=seed_i,
                    base_dir=base_dir,
                    out_dir=out_dir,
                    replica_xlsx=replica_path,
                    python_exe=sys.executable,
                    gurobi_threads=args.gurobi_threads,
                    time_limit=args.time_limit,
                    mip_gap=args.mip_gap,
                    force_base_only=not args.keep_scenarios,
                )
            )

        for fut in as_completed(futures):
            result = fut.result()
            results.append(result)
            print(
                f"[Replica {result.replica}] {result.status} | "
                f"returncode={result.returncode} | "
                f"tiempo={result.elapsed_seconds:.1f}s | "
                f"log={Path(result.log_file).name}"
            )

    results.sort(key=lambda r: r.replica)
    summary_path = out_dir / "resumen_replicas.csv"
    write_summary_csv(summary_path, results)

    n_ok = sum(1 for r in results if r.status == "OK")
    print("\n" + "=" * 72)
    print(f"Finalizado: {n_ok}/{len(results)} replicas OK")
    print(f"Resumen: {summary_path}")
    print("Carpeta con outputs:", out_dir)
    print("Outputs esperados por replica:")
    print("  DatosV2-i.xlsx")
    print("  solution_interday-i.xlsx")
    print("  schedule_resultado-i.csv")
    print("  resumen_diario-i.csv")
    print("  log_interdia-i.txt")
    print("=" * 72)

    if n_ok < len(results):
        print("\nAlgunas replicas fallaron. Revisa los archivos log_interdia-i.txt correspondientes.")
        sys.exit(1)


if __name__ == "__main__":
    main()
