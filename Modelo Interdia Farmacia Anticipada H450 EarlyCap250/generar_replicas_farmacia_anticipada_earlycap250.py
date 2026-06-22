"""
generar_replicas_farmacia_anticipada_earlycap250.py

Genera X replicas de DatosV2.xlsx cambiando SOLO las simulaciones de llegadas
Poisson en la hoja Motor_Arribos, y luego ejecuta model_interdia_farmacia_anticipada_earlycap250.py
para cada replica, dejando outputs atribuibles a cada una en una subcarpeta.

Adaptacion de generar_replicas_v3.py para el modelo INTERDIA FARMACIA ANTICIPADA (H450):
  - Parchea params (no params_v3) en memoria.
  - Lee/escribe solution_interday_farmacia_anticipada_h450_earlycap250-i.xlsx
  - Llama a model_interdia_farmacia_anticipada_earlycap250.main()

Uso tipico desde la carpeta:
    CapstoneG15/Modelo Interdia Farmacia Anticipada H450 EarlyCap250

    python generar_replicas_farmacia_anticipada_earlycap250.py --replicas 30 --workers 2 --gurobi-threads 1 --overwrite

Supuestos:
    - Este archivo esta en el mismo directorio que:
        params.py
        model_interdia_farmacia_anticipada_earlycap250.py
    - El Excel base (DatosV2.xlsx) esta en ../Modelo Interdia/DatosV2.xlsx.
    - Los paths se resuelven buscando en: ./, ../Modelo Interdia/, ../Data Inicial/
    - El horizonte (450) viene del params.py de esta carpeta; el runner no lo toca.
"""

from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook


# Tag usado en los nombres de archivo. Se deriva del OUTPUT_XLSX de params.py
# para mantener consistencia con el modelo, pero se fija aqui para evitar
# acoples con el contenido de params.py.
OUT_TAG = "farmacia_anticipada_h450_earlycap250"


# =============================================================================
# Utilidades de Excel (copiadas de generar_replicas_v3.py)
# =============================================================================

def _header_map(ws, row: int = 1) -> Dict[str, int]:
    out = {}
    for cell in ws[row]:
        if cell.value is not None:
            out[str(cell.value).strip()] = cell.column
    return out


def _get_required(mapping: Dict[str, int], name: str, sheet_name: str) -> int:
    if name not in mapping:
        raise ValueError(
            f"No se encontro la columna '{name}' en la hoja '{sheet_name}'. "
            f"Columnas disponibles: {list(mapping.keys())}"
        )
    return mapping[name]


def leer_tasas_configuracion(xlsx_path: Path) -> Dict[int, float]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    if "Configuracion" not in wb.sheetnames:
        raise ValueError("El archivo no tiene hoja 'Configuracion'.")

    ws = wb["Configuracion"]
    headers = _header_map(ws)
    id_col = _get_required(headers, "Id", "Configuracion")
    tasa_col = _get_required(headers, "Tasa de Llegada", "Configuracion")

    tasas: Dict[int, float] = {}
    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, id_col).value
        if raw_id is None or raw_id == "":
            continue
        try:
            tipo_id = int(raw_id)
        except (TypeError, ValueError):
            continue

        raw_tasa = ws.cell(r, tasa_col).value
        if raw_tasa is None or raw_tasa == "":
            tasa = 0.0
        else:
            tasa = float(raw_tasa)

        if tasa < 0:
            raise ValueError(f"Tasa negativa para tipo {tipo_id}: {tasa}")

        tasas[tipo_id] = tasa

    if not tasas:
        raise ValueError("No se encontraron tasas en Configuracion.")

    wb.close()
    return tasas


def columnas_tipos_motor_arribos(ws) -> List[Tuple[int, int]]:
    out: List[Tuple[int, int]] = []
    for cell in ws[1]:
        value = cell.value
        if not isinstance(value, str):
            continue
        value = value.strip()
        if not value.lower().startswith("tipo"):
            continue
        try:
            tipo_id = int(value.split()[-1])
        except (IndexError, ValueError):
            continue
        out.append((tipo_id, cell.column))
    out.sort(key=lambda x: x[0])
    if not out:
        raise ValueError("No se encontraron columnas 'Tipo X' en Motor_Arribos.")
    return out


def filas_dias_motor_arribos(ws) -> List[int]:
    rows: List[int] = []
    for r in range(2, ws.max_row + 1):
        dia = ws.cell(r, 1).value
        if dia is None or dia == "":
            continue
        try:
            int(dia)
        except (TypeError, ValueError):
            continue
        rows.append(r)
    if not rows:
        raise ValueError("No se encontraron dias validos en Motor_Arribos.")
    return rows


def normalizar_duracion_configuracion_wb(wb) -> int:
    if "Configuracion" not in wb.sheetnames:
        return 0

    ws = wb["Configuracion"]
    headers = _header_map(ws)

    required = ["Id", "Ciclos", "Sesiones", "TBS", "TBC", "Duracion (Dias)"]
    if any(name not in headers for name in required):
        return 0

    id_col = headers["Id"]
    ciclos_col = headers["Ciclos"]
    sesiones_col = headers["Sesiones"]
    tbs_col = headers["TBS"]
    tbc_col = headers["TBC"]
    dur_col = headers["Duracion (Dias)"]

    changed = 0
    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, id_col).value
        if raw_id is None or raw_id == "":
            continue

        try:
            ciclos = int(ws.cell(r, ciclos_col).value)
            sesiones = int(ws.cell(r, sesiones_col).value)
            tbs = int(ws.cell(r, tbs_col).value)
            tbc = int(ws.cell(r, tbc_col).value)
        except (TypeError, ValueError):
            continue

        duracion = (ciclos - 1) * ((sesiones - 1) * tbs + tbc) + (sesiones - 1) * tbs
        ws.cell(r, dur_col).value = int(duracion)
        changed += 1

    return changed


def reparar_duracion_configuracion(xlsx_path: Path) -> int:
    wb = load_workbook(xlsx_path, data_only=False)
    changed = normalizar_duracion_configuracion_wb(wb)
    if changed:
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
        except Exception:
            pass
        wb.save(xlsx_path)
    wb.close()
    return changed


def generar_replica_datosv2(
    base_xlsx: Path,
    out_xlsx: Path,
    seed: int,
    overwrite: bool = False,
) -> Dict[str, object]:
    if out_xlsx.exists() and not overwrite:
        raise FileExistsError(f"Ya existe {out_xlsx}. Usa --overwrite para reemplazarlo.")

    tasas = leer_tasas_configuracion(base_xlsx)

    wb = load_workbook(base_xlsx, data_only=False)
    if "Motor_Arribos" not in wb.sheetnames:
        raise ValueError("El archivo no tiene hoja 'Motor_Arribos'.")

    ws_arribos = wb["Motor_Arribos"]
    tipo_cols = columnas_tipos_motor_arribos(ws_arribos)
    day_rows = filas_dias_motor_arribos(ws_arribos)

    missing_rates = [tipo_id for tipo_id, _ in tipo_cols if tipo_id not in tasas]
    if missing_rates:
        raise ValueError(f"Faltan tasas para tipos: {missing_rates}")

    rng = np.random.default_rng(seed)

    total_llegadas = 0
    for tipo_id, col in tipo_cols:
        lam = tasas[tipo_id]
        values = rng.poisson(lam=lam, size=len(day_rows))
        total_llegadas += int(values.sum())
        for row_idx, value in zip(day_rows, values):
            ws_arribos.cell(row=row_idx, column=col).value = int(value)

    normalizar_duracion_configuracion_wb(wb)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    wb.close()

    return {
        "archivo": str(out_xlsx),
        "seed": seed,
        "n_dias": len(day_rows),
        "n_tipos": len(tipo_cols),
        "total_llegadas_simuladas": total_llegadas,
    }


# =============================================================================
# Ejecucion del modelo interdia Farmacia Anticipada por replica
# =============================================================================

@dataclass
class RunResult:
    replica: int
    seed: int
    datos_xlsx: str
    output_xlsx: str
    log_file: str
    returncode: int
    status: str
    elapsed_seconds: float


def build_model_runner_code(
    base_dir: Path,
    replica_xlsx: Path,
    output_xlsx: str,
    gurobi_threads: Optional[int],
    time_limit: Optional[int],
    mip_gap: Optional[float],
    stall_seconds: int,
    stall_min_runtime: int,
    stall_min_gap_improvement: float,
    disable_gap_stall: bool,
    force_base_only: bool,
) -> str:
    """
    Codigo Python que se ejecuta en un subprocess. Parchea params en memoria.
    No modifica params.py en disco.
    """
    lines = [
        "from pathlib import Path",
        "import sys",
        "try:",
        "    sys.stdout.reconfigure(encoding='utf-8', errors='replace')",
        "    sys.stderr.reconfigure(encoding='utf-8', errors='replace')",
        "except Exception:",
        "    pass",
        f"sys.path.insert(0, {str(base_dir)!r})",
        "import params as P",
        f"P.EXCEL_PATH = Path({str(replica_xlsx)!r})",
        f"P.OUTPUT_XLSX = {output_xlsx!r}",
        f"P.OUTPUT_CSV = ''",
        f"P.OUTPUT_SUMMARY_CSV = ''",
        f"P.USE_GAP_STALL_CALLBACK = {not disable_gap_stall!r}",
        f"P.STALL_SECONDS = {int(stall_seconds)}",
        f"P.STALL_MIN_RUNTIME = {int(stall_min_runtime)}",
        f"P.STALL_MIN_GAP_IMPROVEMENT = {float(stall_min_gap_improvement)!r}",
    ]

    if gurobi_threads is not None:
        lines.append(f"P.THREADS = {int(gurobi_threads)}")
    if time_limit is not None:
        lines.append(f"P.TIME_LIMIT_SECONDS = {int(time_limit)}")
    if mip_gap is not None:
        lines.append(f"P.MIP_GAP = {float(mip_gap)!r}")
    if force_base_only:
        lines.append("P.RUN_ALL_SCENARIOS = False")

    lines.extend([
        "import model_interdia_farmacia_anticipada_earlycap250",
        "model_interdia_farmacia_anticipada_earlycap250.main()",
    ])
    return "\n".join(lines)


def run_model_for_replica(
    replica: int,
    seed: int,
    base_dir: Path,
    out_dir: Path,
    replica_xlsx: Path,
    python_exe: str,
    gurobi_threads: Optional[int],
    time_limit: Optional[int],
    mip_gap: Optional[float],
    stall_seconds: int,
    stall_min_runtime: int,
    stall_min_gap_improvement: float,
    disable_gap_stall: bool,
    force_base_only: bool,
) -> RunResult:
    output_xlsx_path = out_dir / f"solution_interday_{OUT_TAG}-{replica}.xlsx"
    log_file = out_dir / f"log_interdia_{OUT_TAG}-{replica}.txt"

    code = build_model_runner_code(
        base_dir=base_dir.resolve(),
        replica_xlsx=replica_xlsx.resolve(),
        output_xlsx=str(output_xlsx_path.resolve()),
        gurobi_threads=gurobi_threads,
        time_limit=time_limit,
        mip_gap=mip_gap,
        stall_seconds=stall_seconds,
        stall_min_runtime=stall_min_runtime,
        stall_min_gap_improvement=stall_min_gap_improvement,
        disable_gap_stall=disable_gap_stall,
        force_base_only=force_base_only,
    )

    t0 = time.time()
    with open(log_file, "w", encoding="utf-8") as log:
        log.write(f"Replica: {replica}\n")
        log.write(f"Seed: {seed}\n")
        log.write(f"Datos: {replica_xlsx}\n")
        log.write(f"Output XLSX: {output_xlsx_path}\n")
        log.write(f"Modelo: interdia Farmacia Anticipada H450 (prep t-1)\n")
        log.write("\n--- INICIO MODELO FARMACIA ANTICIPADA ---\n\n")
        log.flush()

        env = os.environ.copy()
        env["PYTHONUTF8"] = "1"
        env["PYTHONIOENCODING"] = "utf-8"

        completed = subprocess.run(
            [python_exe, "-X", "utf8", "-c", code],
            cwd=out_dir,
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
        )

    elapsed = time.time() - t0
    status = "OK" if completed.returncode == 0 else "ERROR"

    return RunResult(
        replica=replica,
        seed=seed,
        datos_xlsx=str(replica_xlsx),
        output_xlsx=str(output_xlsx_path),
        log_file=str(log_file),
        returncode=completed.returncode,
        status=status,
        elapsed_seconds=round(elapsed, 2),
    )


# =============================================================================
# Resolucion de rutas automaticas
# =============================================================================

def _find_file(candidates: List[Path], label: str) -> Path:
    for c in candidates:
        if c.exists():
            return c.resolve()
    tried = "\n  - ".join(str(c) for c in candidates)
    raise FileNotFoundError(f"No se encontro {label}.\nRutas probadas:\n  - {tried}")


def resolve_paths(script_dir: Path, args):
    """Resuelve todas las rutas de archivos necesarias."""
    # Excel base: buscar en ../Modelo Interdia/ y ubicaciones comunes
    base_excel = _find_file(
        [
            script_dir / "DatosV2.xlsx",
            script_dir.parent / "Modelo Interdia" / "DatosV2.xlsx",
            script_dir.parent / "Data Inicial" / "DatosV2.xlsx",
            Path.cwd() / "DatosV2.xlsx",
        ],
        "DatosV2.xlsx",
    )

    # Validar que params.py y model_interdia_farmacia_anticipada_earlycap250.py existen en base_dir
    _base_dir_candidate = Path(args.base_dir) if args.base_dir else script_dir
    _params_candidate = _base_dir_candidate / "params.py"
    if not _params_candidate.exists():
        raise FileNotFoundError(
            f"No se encontro params.py en {_base_dir_candidate}"
        )
    base_dir = _params_candidate.parent

    required_files = [
        base_dir / "params.py",
        base_dir / "model_interdia_farmacia_anticipada_earlycap250.py",
    ]
    missing = [str(p) for p in required_files if not p.exists()]
    if missing:
        raise FileNotFoundError(
            f"Faltan archivos requeridos del modelo Farmacia Anticipada:\n  - "
            + "\n  - ".join(missing)
        )

    return base_dir, base_excel


# =============================================================================
# Main
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera replicas Poisson de DatosV2 y ejecuta model_interdia_farmacia_anticipada_earlycap250 para cada replica."
    )
    parser.add_argument("--replicas", type=int, default=30, help="Cantidad de replicas a generar/correr.")
    parser.add_argument("--start-replica", type=int, default=1, help="Numero inicial de replica.")
    parser.add_argument("--base-dir", type=str, default=None, help="Directorio donde estan params.py y model_interdia_farmacia_anticipada_earlycap250.py. Default: script dir.")
    parser.add_argument("--base-excel", type=str, default=None, help="Archivo base Excel. Default: ../Modelo Interdia/DatosV2.xlsx.")
    parser.add_argument("--out-dir", type=str, default=f"resultados_{OUT_TAG}", help="Subcarpeta donde se guardan replicas, outputs, logs y resumen.")
    parser.add_argument("--seed", type=int, default=12345, help="Seed base. La replica i usa seed + i.")
    parser.add_argument("--workers", type=int, default=2, help="Cantidad de replicas a resolver en paralelo.")
    parser.add_argument("--gurobi-threads", type=int, default=1, help="Threads de Gurobi por replica.")
    parser.add_argument("--time-limit", type=int, default=None, help="Sobrescribe P.TIME_LIMIT_SECONDS.")
    parser.add_argument("--mip-gap", type=float, default=None, help="Sobrescribe P.MIP_GAP.")
    parser.add_argument("--stall-seconds", type=int, default=60, help="Segundos sin mejora de gap antes de cortar.")
    parser.add_argument("--stall-min-runtime", type=int, default=120, help="Runtime minimo antes de permitir corte.")
    parser.add_argument("--stall-min-gap-improvement", type=float, default=1e-4, help="Mejora minima de gap.")
    parser.add_argument("--disable-gap-stall", action="store_true", help="Desactiva callback de corte por estancamiento.")
    parser.add_argument("--overwrite", action="store_true", help="Sobrescribe replicas/outputs existentes.")
    parser.add_argument("--solo-generar", action="store_true", help="Solo genera DatosV2-i.xlsx; no corre modelo.")
    parser.add_argument("--solo-correr", action="store_true", help="No genera Excel; usa replicas ya existentes.")
    parser.add_argument("--keep-scenarios", action="store_true", help="No fuerza RUN_ALL_SCENARIOS=False.")
    return parser.parse_args()


def write_summary_csv(path: Path, results: List[RunResult]) -> None:
    if not results:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(results[0]).keys()))
        writer.writeheader()
        for result in results:
            writer.writerow(asdict(result))


def main() -> None:
    args = parse_args()
    if args.replicas <= 0:
        raise ValueError("--replicas debe ser >= 1")
    if args.start_replica <= 0:
        raise ValueError("--start-replica debe ser >= 1")
    if args.workers <= 0:
        raise ValueError("--workers debe ser >= 1")
    if args.stall_seconds <= 0:
        raise ValueError("--stall-seconds debe ser >= 1")
    if args.stall_min_runtime < 0:
        raise ValueError("--stall-min-runtime debe ser >= 0")
    if args.solo_generar and args.solo_correr:
        raise ValueError("No puedes usar --solo-generar y --solo-correr al mismo tiempo.")

    script_dir = Path(__file__).resolve().parent
    base_dir, base_excel = resolve_paths(script_dir, args)

    if args.base_excel:
        base_excel_arg = Path(args.base_excel)
        if not base_excel_arg.is_absolute():
            base_excel_arg = base_dir / base_excel_arg
        base_excel = base_excel_arg.resolve()

    out_dir = Path(args.out_dir)
    if not out_dir.is_absolute():
        out_dir = base_dir / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    end_replica = args.start_replica + args.replicas - 1
    print("=" * 72)
    print("GENERADOR DE REPLICAS + EJECUCION MODEL_INTERDIA FARMACIA ANTICIPADA")
    print("=" * 72)
    print(f"Directorio modelo FA:  {base_dir}")
    print(f"Carpeta salida:        {out_dir}")
    print(f"Excel base:            {base_excel}")
    print(f"Replicas:              {args.replicas} ({args.start_replica}..{end_replica})")
    print(f"Workers modelo:        {args.workers}")
    print(f"Threads/replica:       {args.gurobi_threads}")
    print(f"Gap stall:             {'desactivado' if args.disable_gap_stall else 'activado'}")
    print("=" * 72)

    # -------------------------------------------------------------------------
    # 1) Generar replicas DatosV2-i.xlsx
    # -------------------------------------------------------------------------
    replica_paths: List[Tuple[int, int, Path]] = []
    if not args.solo_correr:
        for i in range(args.start_replica, end_replica + 1):
            seed_i = args.seed + i
            replica_path = out_dir / f"DatosV2-{i}.xlsx"
            info = generar_replica_datosv2(
                base_xlsx=base_excel,
                out_xlsx=replica_path,
                seed=seed_i,
                overwrite=args.overwrite,
            )
            replica_paths.append((i, seed_i, replica_path))
            print(
                f"[Replica {i}] generado {replica_path.name} | "
                f"seed={seed_i} | llegadas={info['total_llegadas_simuladas']}"
            )
    else:
        for i in range(args.start_replica, end_replica + 1):
            seed_i = args.seed + i
            replica_path = out_dir / f"DatosV2-{i}.xlsx"
            if not replica_path.exists():
                raise FileNotFoundError(f"No existe {replica_path}. Quita --solo-correr o genera antes.")
            reparar_duracion_configuracion(replica_path)
            replica_paths.append((i, seed_i, replica_path))

    if args.solo_generar:
        print("\nListo: replicas generadas. No se ejecuto model_interdia_farmacia_anticipada por --solo-generar.")
        return

    # -------------------------------------------------------------------------
    # 2) Ejecutar model_interdia_farmacia_anticipada para cada replica
    # -------------------------------------------------------------------------
    print("\nEjecutando model_interdia_farmacia_anticipada_earlycap250 por replica...")
    results: List[RunResult] = []

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = []
        for i, seed_i, replica_path in replica_paths:
            futures.append(
                executor.submit(
                    run_model_for_replica,
                    replica=i,
                    seed=seed_i,
                    base_dir=base_dir,
                    out_dir=out_dir,
                    replica_xlsx=replica_path,
                    python_exe=sys.executable,
                    gurobi_threads=args.gurobi_threads,
                    time_limit=args.time_limit,
                    mip_gap=args.mip_gap,
                    stall_seconds=args.stall_seconds,
                    stall_min_runtime=args.stall_min_runtime,
                    stall_min_gap_improvement=args.stall_min_gap_improvement,
                    disable_gap_stall=args.disable_gap_stall,
                    force_base_only=not args.keep_scenarios,
                )
            )

        for fut in as_completed(futures):
            result = fut.result()
            results.append(result)
            print(
                f"[Replica {result.replica}] {result.status} | "
                f"returncode={result.returncode} | "
                f"tiempo={result.elapsed_seconds:.1f}s | "
                f"log={Path(result.log_file).name}"
            )

    results.sort(key=lambda r: r.replica)

    # Resumen de corridas (metadata: seed, paths, returncode, tiempo)
    summary_csv = out_dir / f"resumen_corridas_{OUT_TAG}.csv"
    write_summary_csv(summary_csv, results)
    print(f"\nResumen de corridas: {summary_csv}")

    n_ok = sum(1 for r in results if r.status == "OK")
    print("\n" + "=" * 72)
    print(f"Finalizado: {n_ok}/{len(results)} replicas OK")
    print("Carpeta con outputs:", out_dir)
    print("Outputs esperados por replica:")
    print(f"  DatosV2-i.xlsx")
    print(f"  solution_interday_{OUT_TAG}-i.xlsx")
    print(f"  log_interdia_{OUT_TAG}-i.txt")
    print("=" * 72)

    if n_ok < len(results):
        print(f"\nAlgunas replicas fallaron. Revisa los archivos log_interdia_{OUT_TAG}-i.txt correspondientes.")
        sys.exit(1)


if __name__ == "__main__":
    main()
