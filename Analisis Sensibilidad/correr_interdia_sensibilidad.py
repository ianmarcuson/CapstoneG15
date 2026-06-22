from __future__ import annotations

import argparse
import csv
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook

from escenarios_sensibilidad import SensitivityScenario, get_scenarios


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
INTERDIA_DIR = PROJECT_DIR / "Modelo Interdia Farmacia Anticipada H450 EarlyCap250"
MODEL_MODULE = "model_interdia_farmacia_anticipada_earlycap250"
BASE_EXCEL = PROJECT_DIR / "Modelo Interdia" / "DatosV2.xlsx"
OUT_ROOT = SCRIPT_DIR / "resultados_interdia"


@dataclass
class InterdiaSensitivityResult:
    scenario_id: str
    scenario_label: str
    replica: int
    seed: int
    datos_xlsx: str
    output_xlsx: str
    log_file: str
    returncode: int
    status: str
    elapsed_seconds: float


def _header_map(ws, row: int = 1) -> Dict[str, int]:
    return {str(cell.value).strip(): cell.column for cell in ws[row] if cell.value is not None}


def _required(headers: Dict[str, int], candidates: List[str], sheet: str) -> int:
    for name in candidates:
        if name in headers:
            return headers[name]
    raise ValueError(f"No se encontro ninguna columna {candidates} en hoja {sheet}.")


def _type_rates(wb) -> Dict[int, float]:
    ws = wb["Configuracion"]
    headers = _header_map(ws)
    id_col = _required(headers, ["Id"], "Configuracion")
    tasa_col = _required(headers, ["Tasa de Llegada"], "Configuracion")
    rates: Dict[int, float] = {}
    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, id_col).value
        if raw_id in (None, ""):
            continue
        try:
            tipo_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        raw_rate = ws.cell(r, tasa_col).value
        rates[tipo_id] = 0.0 if raw_rate in (None, "") else float(raw_rate)
    return rates


def _motor_columns(ws) -> List[Tuple[int, int]]:
    out: List[Tuple[int, int]] = []
    for cell in ws[1]:
        if not isinstance(cell.value, str):
            continue
        value = cell.value.strip()
        if not value.lower().startswith("tipo"):
            continue
        out.append((int(value.split()[-1]), cell.column))
    out.sort(key=lambda x: x[0])
    return out


def _day_rows(ws) -> List[int]:
    rows: List[int] = []
    for r in range(2, ws.max_row + 1):
        value = ws.cell(r, 1).value
        if value in (None, ""):
            continue
        try:
            int(value)
        except (TypeError, ValueError):
            continue
        rows.append(r)
    return rows


def _normalize_duration(wb) -> None:
    ws = wb["Configuracion"]
    headers = _header_map(ws)
    required = ["Id", "Ciclos", "Sesiones", "TBS", "TBC", "Duracion (Dias)"]
    if any(col not in headers for col in required):
        return
    for r in range(2, ws.max_row + 1):
        raw_id = ws.cell(r, headers["Id"]).value
        if raw_id in (None, ""):
            continue
        try:
            ciclos = int(ws.cell(r, headers["Ciclos"]).value)
            sesiones = int(ws.cell(r, headers["Sesiones"]).value)
            tbs = int(ws.cell(r, headers["TBS"]).value)
            tbc = int(ws.cell(r, headers["TBC"]).value)
        except (TypeError, ValueError):
            continue
        ws.cell(r, headers["Duracion (Dias)"]).value = int(
            (ciclos - 1) * ((sesiones - 1) * tbs + tbc) + (sesiones - 1) * tbs
        )


def _apply_treatment_duration_multiplier(wb, multiplier: float) -> None:
    if abs(multiplier - 1.0) < 1e-12:
        return
    ws = wb["Configuracion"]
    headers = _header_map(ws)
    mod_col = _required(headers, ["Modulos", "Módulos", "MÃ³dulos"], "Configuracion")
    for r in range(2, ws.max_row + 1):
        value = ws.cell(r, mod_col).value
        if value in (None, ""):
            continue
        ws.cell(r, mod_col).value = int(np.ceil(float(value) * multiplier))


def generate_scenario_replica(
    scenario: SensitivityScenario,
    replica: int,
    seed: int,
    out_xlsx: Path,
    overwrite: bool,
) -> int:
    if out_xlsx.exists() and not overwrite:
        raise FileExistsError(f"Ya existe {out_xlsx}. Usa --overwrite para reemplazarlo.")

    wb = load_workbook(BASE_EXCEL, data_only=False)
    rates = _type_rates(wb)
    _apply_treatment_duration_multiplier(wb, scenario.treatment_duration_multiplier)

    ws_arrivals = wb["Motor_Arribos"]
    tipo_cols = _motor_columns(ws_arrivals)
    rows = _day_rows(ws_arrivals)
    rng = np.random.default_rng(seed)
    total_arrivals = 0
    for tipo_id, col in tipo_cols:
        lam = rates[tipo_id] * scenario.arrival_multiplier
        values = rng.poisson(lam=lam, size=len(rows))
        total_arrivals += int(values.sum())
        for row_idx, value in zip(rows, values):
            ws_arrivals.cell(row_idx, col).value = int(value)

    _normalize_duration(wb)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    wb.close()
    return total_arrivals


def build_model_code(
    scenario: SensitivityScenario,
    replica_xlsx: Path,
    output_xlsx: Path,
    gurobi_threads: Optional[int],
    time_limit: Optional[int],
    mip_gap: Optional[float],
    disable_gap_stall: bool,
) -> str:
    lines = [
        "from pathlib import Path",
        "import sys",
        "sys.stdout.reconfigure(encoding='utf-8', errors='replace')",
        "sys.stderr.reconfigure(encoding='utf-8', errors='replace')",
        f"sys.path.insert(0, {str(INTERDIA_DIR)!r})",
        "import params as P",
        f"P.EXCEL_PATH = Path({str(replica_xlsx.resolve())!r})",
        f"P.OUTPUT_XLSX = {str(output_xlsx.resolve())!r}",
        "P.OUTPUT_CSV = ''",
        "P.OUTPUT_SUMMARY_CSV = ''",
        f"P.EARLY_PREP_TREATMENT_CAP = {int(scenario.early_prep_cap)}",
        f"P.USE_GAP_STALL_CALLBACK = {not disable_gap_stall!r}",
        "P.RUN_ALL_SCENARIOS = False",
    ]
    if scenario.n_sillas_override is not None:
        lines.append(f"P.N_SILLAS_OVERRIDE = {int(scenario.n_sillas_override)}")
    if gurobi_threads is not None:
        lines.append(f"P.THREADS = {int(gurobi_threads)}")
    if time_limit is not None:
        lines.append(f"P.TIME_LIMIT_SECONDS = {int(time_limit)}")
    if mip_gap is not None:
        lines.append(f"P.MIP_GAP = {float(mip_gap)!r}")
    lines.extend([f"import {MODEL_MODULE}", f"{MODEL_MODULE}.main()"])
    return "\n".join(lines)


def run_one(
    scenario: SensitivityScenario,
    replica: int,
    seed: int,
    out_dir: Path,
    overwrite: bool,
    gurobi_threads: Optional[int],
    time_limit: Optional[int],
    mip_gap: Optional[float],
    disable_gap_stall: bool,
    prepare_only: bool,
) -> InterdiaSensitivityResult:
    datos_xlsx = out_dir / f"DatosV2-{scenario.id}-{replica}.xlsx"
    output_xlsx = out_dir / f"solution_interday_{scenario.id}-{replica}.xlsx"
    log_file = out_dir / f"log_interdia_{scenario.id}-{replica}.txt"

    total_arrivals = generate_scenario_replica(scenario, replica, seed, datos_xlsx, overwrite)
    if prepare_only:
        return InterdiaSensitivityResult(
            scenario_id=scenario.id,
            scenario_label=scenario.label,
            replica=replica,
            seed=seed,
            datos_xlsx=str(datos_xlsx),
            output_xlsx=str(output_xlsx),
            log_file=str(log_file),
            returncode=0,
            status=f"PREPARED arrivals={total_arrivals}",
            elapsed_seconds=0.0,
        )

    code = build_model_code(scenario, datos_xlsx, output_xlsx, gurobi_threads, time_limit, mip_gap, disable_gap_stall)
    t0 = time.time()
    with open(log_file, "w", encoding="utf-8", errors="replace") as log:
        log.write(f"Scenario: {scenario.id} | {scenario.label}\n")
        log.write(f"Replica: {replica} | seed={seed} | llegadas={total_arrivals}\n")
        log.write(f"Datos: {datos_xlsx}\n")
        log.write(f"Output: {output_xlsx}\n\n")
        log.flush()
        completed = subprocess.run(
            [sys.executable, "-X", "utf8", "-c", code],
            cwd=INTERDIA_DIR,
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
        )
    elapsed = round(time.time() - t0, 2)
    return InterdiaSensitivityResult(
        scenario_id=scenario.id,
        scenario_label=scenario.label,
        replica=replica,
        seed=seed,
        datos_xlsx=str(datos_xlsx),
        output_xlsx=str(output_xlsx),
        log_file=str(log_file),
        returncode=completed.returncode,
        status="OK" if completed.returncode == 0 else "ERROR",
        elapsed_seconds=elapsed,
    )


def write_summary(path: Path, rows: List[InterdiaSensitivityResult]) -> None:
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(rows[0]).keys()))
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Corre escenarios de sensibilidad del modelo interdia.")
    parser.add_argument("--scenarios", nargs="*", default=None)
    parser.add_argument("--replicas", type=int, default=1)
    parser.add_argument("--start-replica", type=int, default=1)
    parser.add_argument("--seed", type=int, default=12345)
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--gurobi-threads", type=int, default=6)
    parser.add_argument("--time-limit", type=int, default=1800)
    parser.add_argument("--mip-gap", type=float, default=0.01)
    parser.add_argument("--disable-gap-stall", action="store_true", default=True)
    parser.add_argument(
        "--prepare-only",
        action="store_true",
        help="Solo genera los DatosV2 modificados y el resumen; no ejecuta Gurobi.",
    )
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    scenarios = [s for s in get_scenarios(args.scenarios) if s.run_interdia]
    end_replica = args.start_replica + args.replicas - 1
    all_results: List[InterdiaSensitivityResult] = []

    futures = []
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        for scenario in scenarios:
            out_dir = OUT_ROOT / scenario.id
            out_dir.mkdir(parents=True, exist_ok=True)
            for replica in range(args.start_replica, end_replica + 1):
                seed = args.seed + replica
                futures.append(
                    executor.submit(
                        run_one,
                        scenario,
                        replica,
                        seed,
                        out_dir,
                        args.overwrite,
                        args.gurobi_threads,
                        args.time_limit,
                        args.mip_gap,
                        args.disable_gap_stall,
                        args.prepare_only,
                    )
                )
        for future in as_completed(futures):
            result = future.result()
            all_results.append(result)
            print(f"[{result.scenario_id} r{result.replica}] {result.status} | {result.elapsed_seconds:.1f}s")

    all_results.sort(key=lambda r: (r.scenario_id, r.replica))
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    write_summary(OUT_ROOT / "resumen_interdia_sensibilidad.csv", all_results)
    if any(r.status == "ERROR" for r in all_results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
